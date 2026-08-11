import argparse
import sys
import time
import unicodedata
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Sequence

import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns
from matplotlib.ticker import FuncFormatter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


ARQUIVO_EXCEL = "DRE 2019.xlsx"
PASTA_GRAFICOS = "graficos_dre_2019"
PASTA_SAIDAS = "saidas_dre_2019"
ARQUIVO_SAIDA = "dre_2019_resultados_executivos.xlsx"


@dataclass(frozen=True)
class ConfigDRE:
    aba_plano_contas: str = "Plano de Contas"
    aba_formato_dre: str = "Formato DRE"
    aba_realizado: str = "Realizado"
    aba_orcado: str = "Orçado"

    coluna_mes_ano: str = "Mês/Ano"
    coluna_conta: str = "Conta"
    coluna_descricao: str = "Descrição da Conta"
    coluna_valor_realizado: str = "Valor Realizado"
    coluna_valor_orcado: str = "Valor Orçado"
    coluna_nivel_1: str = "Nível 1"
    coluna_nivel_2: str = "Nível 2"
    coluna_ordem: str = "Ordem"
    coluna_grupo: str = "Grupo"
    coluna_subtotal: str = "Subtotal?"

    coluna_desvio: str = "Desvio"
    coluna_desvio_percentual: str = "Desvio_%"
    coluna_impacto_resultado: str = "Impacto_Resultado"
    coluna_favorabilidade: str = "Favorabilidade"
    coluna_status: str = "Status"

    limite_alerta_valor: float = 250_000
    limite_critico_valor: float = 1_000_000
    limite_alerta_pct_receita_liquida: float = 0.15
    limite_critico_pct_receita_liquida: float = 0.75
    limite_alerta_pct_lucro_orcado: float = 0.25
    limite_critico_pct_lucro_orcado: float = 1.00


class AnaliseDRESenior:
    indicadores_percentuais = {
        "Margem Bruta",
        "Margem Líquida",
        "Custo / Receita Líquida",
        "Despesas / Receita Líquida",
    }

    cor_favoravel = "#2E7D32"
    cor_desfavoravel = "#C62828"
    cor_neutra = "#546E7A"
    cor_orcado = "#607D8B"
    cor_realizado = "#1565C0"

    def __init__(
        self,
        arquivo_excel: str = ARQUIVO_EXCEL,
        pasta_graficos: str = PASTA_GRAFICOS,
        pasta_saidas: str = PASTA_SAIDAS,
        config: ConfigDRE | None = None,
    ) -> None:
        self.config = config or ConfigDRE()
        self.arquivo_excel = Path(arquivo_excel)
        self.pasta_graficos = Path(pasta_graficos)
        self.pasta_saidas = Path(pasta_saidas)
        self.inicio = time.perf_counter()

        self.df_plano_contas: pd.DataFrame = pd.DataFrame()
        self.df_formato_dre: pd.DataFrame = pd.DataFrame()
        self.df_realizado: pd.DataFrame = pd.DataFrame()
        self.df_orcado: pd.DataFrame = pd.DataFrame()
        self.df_final: pd.DataFrame = pd.DataFrame()
        self.df_auditoria: pd.DataFrame = pd.DataFrame()

    @staticmethod
    def normalizar_texto(valor: object) -> str:
        texto = str(valor).strip().lower()
        texto = unicodedata.normalize("NFKD", texto)
        texto = "".join(caractere for caractere in texto if not unicodedata.combining(caractere))
        return " ".join(texto.split())

    @staticmethod
    def moeda(valor: float, compacto: bool = False) -> str:
        if pd.isna(valor):
            return "-"
        valor_float = float(valor)
        sinal = "-" if valor_float < 0 else ""
        valor_abs = abs(valor_float)

        if compacto:
            if valor_abs >= 1_000_000:
                texto = f"{valor_abs / 1_000_000:,.1f} MM"
            elif valor_abs >= 1_000:
                texto = f"{valor_abs / 1_000:,.1f} mil"
            else:
                texto = f"{valor_abs:,.0f}"
            texto = texto.replace(",", "X").replace(".", ",").replace("X", ".")
            return f"{sinal}R$ {texto}"

        texto = f"{valor_abs:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"{sinal}R$ {texto}"

    @staticmethod
    def percentual(valor: float, casas: int = 2, pp: bool = False) -> str:
        if pd.isna(valor):
            return "-"
        sufixo = " p.p." if pp else "%"
        return f"{float(valor):.{casas}f}".replace(".", ",") + sufixo

    @staticmethod
    def linha(char: str = "=", tamanho: int = 112) -> None:
        print(char * tamanho)

    @staticmethod
    def titulo(texto: str) -> None:
        print()
        AnaliseDRESenior.linha("=")
        print(texto)
        AnaliseDRESenior.linha("=")

    @staticmethod
    def subtitulo(texto: str) -> None:
        print()
        print(texto)
        AnaliseDRESenior.linha("-", min(112, max(60, len(texto) + 20)))

    @staticmethod
    def formatar_milhoes_eixo(valor: float, _pos: int | None = None) -> str:
        if pd.isna(valor):
            return ""
        valor_abs = abs(float(valor))
        sinal = "-" if valor < 0 else ""
        if valor_abs >= 1_000_000:
            texto = f"{valor_abs / 1_000_000:.1f}".replace(".", ",")
            return f"{sinal}R$ {texto} MM"
        texto = f"{valor_abs / 1_000:.0f}".replace(".", ",")
        return f"{sinal}R$ {texto} mil"

    @staticmethod
    def limpar_nome_grupo(grupo: object) -> str:
        texto = str(grupo).strip()
        if ")" in texto[:4]:
            texto = texto.split(")", 1)[1].strip()
        return texto.title() if texto.isupper() else texto

    @staticmethod
    def percentual_seguro(numerador: float, denominador: float) -> float:
        if pd.isna(denominador) or float(denominador) == 0:
            return pd.NA
        return float(numerador) / float(denominador) * 100

    def validar_arquivo(self) -> None:
        if not self.arquivo_excel.exists():
            raise FileNotFoundError(
                f"Arquivo não encontrado: {self.arquivo_excel.resolve()}\n"
                f"Coloque '{ARQUIVO_EXCEL}' na mesma pasta do script ou use --arquivo."
            )

    def carregar_abas(self) -> None:
        self.validar_arquivo()

        self.df_plano_contas = pd.read_excel(self.arquivo_excel, sheet_name=self.config.aba_plano_contas)
        self.df_formato_dre = pd.read_excel(self.arquivo_excel, sheet_name=self.config.aba_formato_dre)
        self.df_realizado = pd.read_excel(self.arquivo_excel, sheet_name=self.config.aba_realizado)
        self.df_orcado = pd.read_excel(self.arquivo_excel, sheet_name=self.config.aba_orcado)

        for df in [self.df_plano_contas, self.df_formato_dre, self.df_realizado, self.df_orcado]:
            df.columns = [str(coluna).strip() for coluna in df.columns]

    def validar_colunas_obrigatorias(self) -> None:
        c = self.config
        regras = {
            c.aba_plano_contas: (self.df_plano_contas, [c.coluna_conta, c.coluna_descricao, c.coluna_nivel_1, c.coluna_nivel_2]),
            c.aba_realizado: (self.df_realizado, [c.coluna_mes_ano, c.coluna_conta, c.coluna_valor_realizado]),
            c.aba_orcado: (self.df_orcado, [c.coluna_mes_ano, c.coluna_conta, c.coluna_valor_orcado]),
            c.aba_formato_dre: (self.df_formato_dre, [c.coluna_ordem, c.coluna_grupo, c.coluna_subtotal]),
        }

        erros: List[str] = []
        for nome_base, (df, colunas) in regras.items():
            faltantes = [coluna for coluna in colunas if coluna not in df.columns]
            if faltantes:
                erros.append(f"{nome_base}: faltam {faltantes}. Colunas disponíveis: {list(df.columns)}")

        if erros:
            raise KeyError("\n".join(erros))

    def limpar_dados(self) -> None:
        c = self.config

        for df in [self.df_realizado, self.df_orcado]:
            df[c.coluna_mes_ano] = pd.to_datetime(df[c.coluna_mes_ano], errors="coerce")
            df[c.coluna_conta] = df[c.coluna_conta].astype(str).str.strip()

        self.df_realizado[c.coluna_valor_realizado] = pd.to_numeric(self.df_realizado[c.coluna_valor_realizado], errors="coerce")
        self.df_orcado[c.coluna_valor_orcado] = pd.to_numeric(self.df_orcado[c.coluna_valor_orcado], errors="coerce")

        for coluna in [c.coluna_conta, c.coluna_nivel_1, c.coluna_nivel_2, c.coluna_descricao]:
            self.df_plano_contas[coluna] = self.df_plano_contas[coluna].astype(str).str.strip()

        self.df_formato_dre[c.coluna_ordem] = pd.to_numeric(self.df_formato_dre[c.coluna_ordem], errors="coerce")
        self.df_formato_dre[c.coluna_grupo] = self.df_formato_dre[c.coluna_grupo].astype(str).str.strip()
        self.df_formato_dre[c.coluna_subtotal] = pd.to_numeric(self.df_formato_dre[c.coluna_subtotal], errors="coerce").fillna(0).astype(int)
        self.df_formato_dre = self.df_formato_dre.sort_values(c.coluna_ordem).reset_index(drop=True)

    def grupos_operacionais_dre(self) -> List[str]:
        c = self.config
        return (
            self.df_formato_dre.loc[self.df_formato_dre[c.coluna_subtotal].eq(0), c.coluna_grupo]
            .dropna()
            .astype(str)
            .tolist()
        )

    def grupo_receita_bruta(self) -> str:
        grupos = self.grupos_operacionais_dre()
        if not grupos:
            raise ValueError("A aba Formato DRE não possui grupos operacionais.")
        return grupos[0]

    def indicador_lucro_liquido(self) -> str:
        c = self.config
        subtotais = self.df_formato_dre.loc[self.df_formato_dre[c.coluna_subtotal].eq(1), c.coluna_grupo]
        if subtotais.empty:
            raise ValueError("A aba Formato DRE precisa ter ao menos um subtotal.")
        return self.limpar_nome_grupo(subtotais.iloc[-1])

    def registrar_auditoria(self, checagens: List[Dict[str, object]]) -> None:
        self.df_auditoria = pd.DataFrame(checagens)

    def validar_integridade_pre_merge(self) -> None:
        c = self.config
        erros: List[str] = []
        checagens: List[Dict[str, object]] = []

        def registrar(nome: str, quantidade: int, detalhe: str, erro: bool = False) -> None:
            status = "ERRO" if erro and quantidade else "OK"
            checagens.append({"Checagem": nome, "Status": status, "Quantidade": quantidade, "Detalhe": detalhe})
            if erro and quantidade:
                erros.append(f"{nome}: {detalhe}")

        dup_real = int(self.df_realizado.duplicated([c.coluna_mes_ano, c.coluna_conta]).sum())
        dup_orc = int(self.df_orcado.duplicated([c.coluna_mes_ano, c.coluna_conta]).sum())
        dup_plano = int(self.df_plano_contas.duplicated([c.coluna_conta]).sum())
        registrar("Duplicidade Realizado por Mês/Ano + Conta", dup_real, f"{dup_real} chaves duplicadas no Realizado", True)
        registrar("Duplicidade Orçado por Mês/Ano + Conta", dup_orc, f"{dup_orc} chaves duplicadas no Orçado", True)
        registrar("Duplicidade Plano de Contas", dup_plano, f"{dup_plano} contas duplicadas no plano", True)

        data_real = int(self.df_realizado[c.coluna_mes_ano].isna().sum())
        data_orc = int(self.df_orcado[c.coluna_mes_ano].isna().sum())
        valor_real = int(self.df_realizado[c.coluna_valor_realizado].isna().sum())
        valor_orc = int(self.df_orcado[c.coluna_valor_orcado].isna().sum())
        registrar("Datas inválidas no Realizado", data_real, f"{data_real} datas inválidas", True)
        registrar("Datas inválidas no Orçado", data_orc, f"{data_orc} datas inválidas", True)
        registrar("Valores nulos no Realizado", valor_real, f"{valor_real} valores nulos", True)
        registrar("Valores nulos no Orçado", valor_orc, f"{valor_orc} valores nulos", True)

        orc_zero = int(self.df_orcado[c.coluna_valor_orcado].eq(0).sum())
        registrar("Orçamento zero", orc_zero, f"{orc_zero} linhas com orçamento zero; percentual de desvio ficaria indefinido", True)

        chaves_real = set(zip(self.df_realizado[c.coluna_mes_ano], self.df_realizado[c.coluna_conta]))
        chaves_orc = set(zip(self.df_orcado[c.coluna_mes_ano], self.df_orcado[c.coluna_conta]))
        sem_orc = chaves_real - chaves_orc
        sem_real = chaves_orc - chaves_real
        registrar("Realizado sem orçamento", len(sem_orc), f"{len(sem_orc)} chaves do Realizado sem orçamento correspondente", True)
        registrar("Orçado sem realizado", len(sem_real), f"{len(sem_real)} chaves do Orçado sem realizado correspondente", True)

        contas_plano = set(self.df_plano_contas[c.coluna_conta].dropna().astype(str))
        contas_real = set(self.df_realizado[c.coluna_conta].dropna().astype(str))
        contas_orc = set(self.df_orcado[c.coluna_conta].dropna().astype(str))
        contas_sem_plano = (contas_real | contas_orc) - contas_plano
        registrar("Contas fora do Plano de Contas", len(contas_sem_plano), f"{len(contas_sem_plano)} contas sem mapeamento no plano", True)

        grupos_formato = set(self.grupos_operacionais_dre())
        grupos_plano = set(self.df_plano_contas[c.coluna_nivel_1].dropna().astype(str))
        grupos_sem_formato = grupos_plano - grupos_formato
        registrar("Grupos fora do Formato DRE", len(grupos_sem_formato), f"{len(grupos_sem_formato)} grupos do plano fora da estrutura oficial da DRE", True)

        linhas_esperadas = len(self.df_realizado)
        linhas_orcado = len(self.df_orcado)
        registrar("Volume Realizado", linhas_esperadas, f"{linhas_esperadas} linhas carregadas", False)
        registrar("Volume Orçado", linhas_orcado, f"{linhas_orcado} linhas carregadas", False)
        registrar("Contas no Plano", len(contas_plano), f"{len(contas_plano)} contas mapeadas", False)

        self.registrar_auditoria(checagens)
        if erros:
            raise ValueError("Validação de integridade falhou:\n- " + "\n- ".join(erros))

    def consolidar_bases(self) -> None:
        c = self.config
        self.validar_integridade_pre_merge()

        df_consolidado = pd.merge(
            self.df_realizado,
            self.df_orcado,
            on=[c.coluna_mes_ano, c.coluna_conta],
            how="inner",
            validate="one_to_one",
        )

        self.df_final = pd.merge(
            df_consolidado,
            self.df_plano_contas[[c.coluna_conta, c.coluna_descricao, c.coluna_nivel_1, c.coluna_nivel_2]],
            on=c.coluna_conta,
            how="left",
            validate="many_to_one",
        )

        self.df_final[c.coluna_desvio] = self.df_final[c.coluna_valor_realizado] - self.df_final[c.coluna_valor_orcado]
        self.df_final[c.coluna_desvio_percentual] = [
            self.percentual_seguro(desvio, orcado)
            for desvio, orcado in zip(self.df_final[c.coluna_desvio], self.df_final[c.coluna_valor_orcado])
        ]
        self.df_final[c.coluna_impacto_resultado] = self.df_final[c.coluna_desvio]
        self.df_final[c.coluna_favorabilidade] = self.df_final[c.coluna_impacto_resultado].apply(self.favorabilidade)
        self.df_final["Ano"] = self.df_final[c.coluna_mes_ano].dt.year
        self.df_final["Mes_Numero"] = self.df_final[c.coluna_mes_ano].dt.month
        self.df_final["Mês"] = self.df_final[c.coluna_mes_ano].dt.strftime("%m/%Y")

    def grupo_valor(self, grupo: str, coluna: str) -> float:
        c = self.config
        serie = self.df_final.loc[self.df_final[c.coluna_nivel_1] == grupo, coluna]
        return float(serie.sum())

    def totais_por_grupo(self, base: pd.DataFrame | None = None) -> Dict[str, Dict[str, float]]:
        c = self.config
        df = self.df_final if base is None else base
        totais = (
            df.groupby(c.coluna_nivel_1, dropna=False)
            .agg(
                Realizado=(c.coluna_valor_realizado, "sum"),
                Orçado=(c.coluna_valor_orcado, "sum"),
            )
            .to_dict("index")
        )
        return {
            str(grupo): {"Realizado": float(valores["Realizado"]), "Orçado": float(valores["Orçado"])}
            for grupo, valores in totais.items()
        }

    def calcular_linhas_dre(self, totais: Dict[str, Dict[str, float]]) -> pd.DataFrame:
        c = self.config
        linhas: List[Dict[str, object]] = []
        acumulado_real = 0.0
        acumulado_orc = 0.0

        for _, linha in self.df_formato_dre.sort_values(c.coluna_ordem).iterrows():
            grupo = str(linha[c.coluna_grupo]).strip()
            subtotal = int(linha[c.coluna_subtotal]) == 1
            if subtotal:
                realizado = acumulado_real
                orcado = acumulado_orc
            else:
                realizado = float(totais.get(grupo, {}).get("Realizado", 0.0))
                orcado = float(totais.get(grupo, {}).get("Orçado", 0.0))
                acumulado_real += realizado
                acumulado_orc += orcado

            desvio = realizado - orcado
            linhas.append(
                {
                    c.coluna_ordem: int(linha[c.coluna_ordem]),
                    "Tipo_Linha": "Subtotal" if subtotal else "Grupo",
                    c.coluna_grupo: grupo,
                    "Indicador": self.limpar_nome_grupo(grupo),
                    "Realizado": realizado,
                    "Orçado": orcado,
                    c.coluna_desvio: desvio,
                    c.coluna_desvio_percentual: self.percentual_seguro(desvio, orcado),
                    c.coluna_impacto_resultado: desvio,
                }
            )

        return self.adicionar_metricas_materialidade(pd.DataFrame(linhas))

    def calcular_dre_estruturada(self) -> pd.DataFrame:
        return self.calcular_linhas_dre(self.totais_por_grupo())

    def materialidade_bases(self) -> Dict[str, float]:
        c = self.config
        totais = self.totais_por_grupo()
        acumulado_orcado = 0.0
        bases_por_subtotal: Dict[str, float] = {}

        for _, linha in self.df_formato_dre.sort_values(c.coluna_ordem).iterrows():
            grupo = str(linha[c.coluna_grupo]).strip()
            subtotal = int(linha[c.coluna_subtotal]) == 1
            if subtotal:
                bases_por_subtotal[self.limpar_nome_grupo(grupo)] = acumulado_orcado
            else:
                acumulado_orcado += float(totais.get(grupo, {}).get("Orçado", 0.0))

        receita_liquida = bases_por_subtotal.get("Receita Líquida", acumulado_orcado)
        lucro_liquido = bases_por_subtotal.get(self.indicador_lucro_liquido(), receita_liquida)
        return {"Receita_Líquida_Orçada": receita_liquida, "Lucro_Líquido_Orçado": lucro_liquido}

    @staticmethod
    def valor_indicador(df: pd.DataFrame, indicador: str, coluna: str) -> float:
        linha = df.loc[df["Indicador"] == indicador, coluna]
        if linha.empty:
            return 0.0
        return float(linha.iloc[0])

    def favorabilidade(self, impacto: float) -> str:
        if pd.isna(impacto) or abs(float(impacto)) < 0.01:
            return "Neutro"
        return "Favorável" if float(impacto) > 0 else "Desfavorável"

    def classificar_materialidade(self, impacto: float, pct_rl: float, pct_lucro: float) -> str:
        c = self.config
        valor = abs(float(impacto)) if not pd.isna(impacto) else 0.0
        pct_rl = abs(float(pct_rl)) if not pd.isna(pct_rl) else 0.0
        pct_lucro = abs(float(pct_lucro)) if not pd.isna(pct_lucro) else 0.0

        if (
            valor >= c.limite_critico_valor
            or pct_rl >= c.limite_critico_pct_receita_liquida
            or pct_lucro >= c.limite_critico_pct_lucro_orcado
        ):
            return "CRÍTICO"
        if (
            valor >= c.limite_alerta_valor
            or pct_rl >= c.limite_alerta_pct_receita_liquida
            or pct_lucro >= c.limite_alerta_pct_lucro_orcado
        ):
            return "ATENÇÃO"
        return "OK"

    def adicionar_metricas_materialidade(self, df: pd.DataFrame) -> pd.DataFrame:
        c = self.config
        tabela = df.copy()

        bases = self.materialidade_bases()
        base_rl = bases["Receita_Líquida_Orçada"]
        base_lucro = bases["Lucro_Líquido_Orçado"] or base_rl

        tabela["Impacto_Absoluto"] = tabela[c.coluna_impacto_resultado].abs()
        tabela["Materialidade_%_RL"] = [
            self.percentual_seguro(abs(valor), abs(base_rl)) for valor in tabela[c.coluna_impacto_resultado]
        ]
        tabela["Materialidade_%_Lucro_Orçado"] = [
            self.percentual_seguro(abs(valor), abs(base_lucro)) for valor in tabela[c.coluna_impacto_resultado]
        ]
        tabela[c.coluna_favorabilidade] = tabela[c.coluna_impacto_resultado].apply(self.favorabilidade)
        tabela[c.coluna_status] = [
            self.classificar_materialidade(impacto, pct_rl, pct_lucro)
            for impacto, pct_rl, pct_lucro in zip(
                tabela[c.coluna_impacto_resultado],
                tabela["Materialidade_%_RL"],
                tabela["Materialidade_%_Lucro_Orçado"],
            )
        ]
        return tabela

    def calcular_kpis(self) -> pd.DataFrame:
        dre = self.calcular_dre_estruturada()
        indicadores = [
            "Receita Bruta",
            "Deduções da Receita",
            "Receita Líquida",
            "Custos de Vendas",
            "Lucro Bruto",
            "Despesas Gerais",
            "Impostos",
            "Lucro Líquido",
        ]
        linhas: List[Dict[str, object]] = []

        for indicador in indicadores:
            linha = dre.loc[dre["Indicador"] == indicador]
            if linha.empty:
                continue
            registro = linha.iloc[0].to_dict()
            linhas.append(
                {
                    "Indicador": indicador,
                    "Realizado": registro["Realizado"],
                    "Orçado": registro["Orçado"],
                    "Desvio": registro["Desvio"],
                    "Desvio_%": registro["Desvio_%"],
                    "Impacto_Resultado": registro["Impacto_Resultado"],
                    "Favorabilidade": registro["Favorabilidade"],
                    "Status": registro["Status"],
                }
            )

        receita_liquida_real = self.valor_indicador(dre, "Receita Líquida", "Realizado")
        receita_liquida_orc = self.valor_indicador(dre, "Receita Líquida", "Orçado")
        lucro_bruto_real = self.valor_indicador(dre, "Lucro Bruto", "Realizado")
        lucro_bruto_orc = self.valor_indicador(dre, "Lucro Bruto", "Orçado")
        lucro_liquido_real = self.valor_indicador(dre, "Lucro Líquido", "Realizado")
        lucro_liquido_orc = self.valor_indicador(dre, "Lucro Líquido", "Orçado")
        custos_real = self.valor_indicador(dre, "Custos de Vendas", "Realizado")
        custos_orc = self.valor_indicador(dre, "Custos de Vendas", "Orçado")
        despesas_real = self.valor_indicador(dre, "Despesas Gerais", "Realizado")
        despesas_orc = self.valor_indicador(dre, "Despesas Gerais", "Orçado")

        ratios = {
            "Margem Bruta": (
                lucro_bruto_real / receita_liquida_real if receita_liquida_real else pd.NA,
                lucro_bruto_orc / receita_liquida_orc if receita_liquida_orc else pd.NA,
            ),
            "Margem Líquida": (
                lucro_liquido_real / receita_liquida_real if receita_liquida_real else pd.NA,
                lucro_liquido_orc / receita_liquida_orc if receita_liquida_orc else pd.NA,
            ),
            "Custo / Receita Líquida": (
                abs(custos_real) / receita_liquida_real if receita_liquida_real else pd.NA,
                abs(custos_orc) / receita_liquida_orc if receita_liquida_orc else pd.NA,
            ),
            "Despesas / Receita Líquida": (
                abs(despesas_real) / receita_liquida_real if receita_liquida_real else pd.NA,
                abs(despesas_orc) / receita_liquida_orc if receita_liquida_orc else pd.NA,
            ),
        }

        for indicador, (realizado, orcado) in ratios.items():
            desvio = realizado - orcado if not pd.isna(realizado) and not pd.isna(orcado) else pd.NA
            linhas.append(
                {
                    "Indicador": indicador,
                    "Realizado": realizado,
                    "Orçado": orcado,
                    "Desvio": desvio,
                    "Desvio_%": self.percentual_seguro(desvio, orcado) if not pd.isna(desvio) else pd.NA,
                    "Impacto_Resultado": pd.NA,
                    "Favorabilidade": "-",
                    "Status": "-",
                }
            )

        return pd.DataFrame(linhas)

    def resumo_nivel_1(self) -> pd.DataFrame:
        c = self.config
        resumo = (
            self.df_final.groupby(c.coluna_nivel_1, dropna=False)
            .agg(
                Realizado=(c.coluna_valor_realizado, "sum"),
                Orçado=(c.coluna_valor_orcado, "sum"),
                Desvio=(c.coluna_desvio, "sum"),
                Contas=(c.coluna_conta, "nunique"),
            )
            .reset_index()
        )
        resumo[c.coluna_desvio_percentual] = [
            self.percentual_seguro(desvio, orcado) for desvio, orcado in zip(resumo["Desvio"], resumo["Orçado"])
        ]
        resumo[c.coluna_impacto_resultado] = resumo["Desvio"]
        ordem = self.df_formato_dre.set_index(c.coluna_grupo)[c.coluna_ordem].to_dict()
        resumo[c.coluna_ordem] = resumo[c.coluna_nivel_1].map(ordem)
        return self.adicionar_metricas_materialidade(resumo).sort_values(c.coluna_ordem)

    def resumo_nivel_2(self) -> pd.DataFrame:
        c = self.config
        resumo = (
            self.df_final.groupby([c.coluna_nivel_1, c.coluna_nivel_2], dropna=False)
            .agg(
                Realizado=(c.coluna_valor_realizado, "sum"),
                Orçado=(c.coluna_valor_orcado, "sum"),
                Desvio=(c.coluna_desvio, "sum"),
                Contas=(c.coluna_conta, "nunique"),
            )
            .reset_index()
        )
        resumo[c.coluna_desvio_percentual] = [
            self.percentual_seguro(desvio, orcado) for desvio, orcado in zip(resumo["Desvio"], resumo["Orçado"])
        ]
        resumo[c.coluna_impacto_resultado] = resumo["Desvio"]
        ordem = self.df_formato_dre.set_index(c.coluna_grupo)[c.coluna_ordem].to_dict()
        resumo[c.coluna_ordem] = resumo[c.coluna_nivel_1].map(ordem)
        return self.adicionar_metricas_materialidade(resumo).sort_values("Impacto_Absoluto", ascending=False)

    def resumo_receita_mensal(self) -> pd.DataFrame:
        c = self.config
        receita = self.df_final[self.df_final[c.coluna_nivel_1] == self.grupo_receita_bruta()].copy()

        mensal = (
            receita.groupby(c.coluna_mes_ano)
            .agg(
                Receita_Realizada=(c.coluna_valor_realizado, "sum"),
                Receita_Orçada=(c.coluna_valor_orcado, "sum"),
            )
            .reset_index()
            .sort_values(c.coluna_mes_ano)
        )

        mensal["Desvio"] = mensal["Receita_Realizada"] - mensal["Receita_Orçada"]
        mensal["Desvio_%"] = [
            self.percentual_seguro(desvio, orcado) for desvio, orcado in zip(mensal["Desvio"], mensal["Receita_Orçada"])
        ]
        mensal["Impacto_Resultado"] = mensal["Desvio"]
        mensal["Desvio_Acumulado"] = mensal["Desvio"].cumsum()
        mensal["Meta_Batida"] = mensal["Receita_Realizada"] >= mensal["Receita_Orçada"]
        mensal["Mês"] = mensal[c.coluna_mes_ano].dt.strftime("%m/%Y")
        return self.adicionar_metricas_materialidade(mensal)

    def resumo_lucro_liquido_mensal(self) -> pd.DataFrame:
        c = self.config
        registros: List[Dict[str, object]] = []
        indicador_lucro = self.indicador_lucro_liquido()

        for mes, base_mes in self.df_final.groupby(c.coluna_mes_ano):
            dre_mes = self.calcular_linhas_dre(self.totais_por_grupo(base_mes))
            lucro_mes = dre_mes.loc[dre_mes["Indicador"] == indicador_lucro].iloc[0]
            realizado = float(lucro_mes["Realizado"])
            orcado = float(lucro_mes["Orçado"])
            desvio = realizado - orcado
            registros.append(
                {
                    c.coluna_mes_ano: mes,
                    "Mês": mes.strftime("%m/%Y"),
                    "Lucro_Realizado": realizado,
                    "Lucro_Orçado": orcado,
                    "Desvio": desvio,
                    "Desvio_%": self.percentual_seguro(desvio, orcado),
                    "Impacto_Resultado": desvio,
                    "Meta_Batida": realizado >= orcado,
                }
            )

        mensal = pd.DataFrame(registros).sort_values(c.coluna_mes_ano)
        mensal["Desvio_Acumulado"] = mensal["Desvio"].cumsum()
        return self.adicionar_metricas_materialidade(mensal)

    def contribuicao_gap_lucro(self) -> pd.DataFrame:
        nivel_1 = self.resumo_nivel_1()
        grupos = self.grupos_operacionais_dre()
        base = nivel_1[nivel_1[self.config.coluna_nivel_1].isin(grupos)].copy()
        gap_lucro = self.valor_indicador(self.calcular_dre_estruturada(), self.indicador_lucro_liquido(), "Desvio")
        base["Participacao_Gap_Lucro_%"] = [
            self.percentual_seguro(impacto, gap_lucro) for impacto in base["Impacto_Resultado"]
        ]
        base["Categoria"] = base[self.config.coluna_nivel_1].apply(self.limpar_nome_grupo)
        return base.sort_values(self.config.coluna_ordem)

    def pareto_desvios(self) -> pd.DataFrame:
        top = self.resumo_nivel_2().copy().sort_values("Impacto_Absoluto", ascending=False)
        total = float(top["Impacto_Absoluto"].sum())
        top["Impacto_Acumulado"] = top["Impacto_Absoluto"].cumsum()
        top["Pareto_%"] = top["Impacto_Acumulado"] / total * 100 if total else pd.NA
        return top

    def melhores_piores_meses(self) -> pd.DataFrame:
        lucro = self.resumo_lucro_liquido_mensal()
        melhores = lucro.nlargest(3, "Desvio").copy()
        piores = lucro.nsmallest(3, "Desvio").copy()
        melhores["Tipo"] = "Melhores meses"
        piores["Tipo"] = "Piores meses"
        return pd.concat([melhores, piores], ignore_index=True)

    def resumo_executivo_df(self) -> pd.DataFrame:
        kpis = self.calcular_kpis()
        lucro = self.resumo_lucro_liquido_mensal()
        receita = self.resumo_receita_mensal()

        def valor(indicador: str, coluna: str) -> float:
            return float(kpis.loc[kpis["Indicador"] == indicador, coluna].iloc[0])

        top_driver = self.resumo_nivel_2().iloc[0]
        linhas = [
            {"Indicador": "Receita Bruta Realizada", "Valor": valor("Receita Bruta", "Realizado"), "Leitura": "Volume anual realizado"},
            {"Indicador": "Receita Bruta Orçada", "Valor": valor("Receita Bruta", "Orçado"), "Leitura": "Base de comparação do orçamento"},
            {"Indicador": "Desvio Receita Bruta", "Valor": valor("Receita Bruta", "Desvio"), "Leitura": "Impacto direto de top line"},
            {"Indicador": "Lucro Líquido Realizado", "Valor": valor("Lucro Líquido", "Realizado"), "Leitura": "Resultado final realizado"},
            {"Indicador": "Lucro Líquido Orçado", "Valor": valor("Lucro Líquido", "Orçado"), "Leitura": "Resultado final esperado"},
            {"Indicador": "Gap de Lucro Líquido", "Valor": valor("Lucro Líquido", "Desvio"), "Leitura": "Diferença realizada contra orçamento"},
            {"Indicador": "Margem Líquida Realizada", "Valor": valor("Margem Líquida", "Realizado"), "Leitura": "Percentual da receita líquida convertido em lucro"},
            {"Indicador": "Meses com Receita acima da meta", "Valor": int(receita["Meta_Batida"].sum()), "Leitura": f"{int(receita['Meta_Batida'].sum())}/{len(receita)} meses"},
            {"Indicador": "Meses com Lucro acima da meta", "Valor": int(lucro["Meta_Batida"].sum()), "Leitura": f"{int(lucro['Meta_Batida'].sum())}/{len(lucro)} meses"},
            {"Indicador": "Maior driver de desvio", "Valor": float(top_driver["Impacto_Resultado"]), "Leitura": str(top_driver[self.config.coluna_nivel_2])},
        ]
        return pd.DataFrame(linhas)

    def auditoria_dados(self) -> pd.DataFrame:
        c = self.config
        linhas = [] if self.df_auditoria.empty else self.df_auditoria.to_dict("records")
        linhas.extend(
            [
                {"Checagem": "Linhas consolidadas", "Status": "OK", "Quantidade": len(self.df_final), "Detalhe": "Linhas após merge Realizado x Orçado x Plano"},
                {"Checagem": "Meses analisados", "Status": "OK", "Quantidade": self.df_final[c.coluna_mes_ano].nunique(), "Detalhe": "Competências únicas na base final"},
                {"Checagem": "Contas analisadas", "Status": "OK", "Quantidade": self.df_final[c.coluna_conta].nunique(), "Detalhe": "Contas únicas na base final"},
                {"Checagem": "Valores nulos base final", "Status": "OK", "Quantidade": int(self.df_final.isna().sum().sum()), "Detalhe": "Total de nulos após consolidação"},
            ]
        )
        return pd.DataFrame(linhas)

    def formatar_tabela_moeda(self, df: pd.DataFrame, colunas_moeda: Iterable[str], colunas_percentuais: Iterable[str] = ()) -> pd.DataFrame:
        tabela = df.copy()
        for coluna in colunas_moeda:
            if coluna in tabela.columns:
                tabela[coluna] = tabela[coluna].apply(self.moeda)
        for coluna in colunas_percentuais:
            if coluna in tabela.columns:
                tabela[coluna] = tabela[coluna].apply(lambda valor: self.percentual(valor))
        return tabela

    def mostrar_resumo_executivo(self) -> None:
        kpis = self.calcular_kpis()
        receita = self.resumo_receita_mensal()
        lucro_mensal = self.resumo_lucro_liquido_mensal()
        top_desvios = self.resumo_nivel_2().head(5)

        def valor(indicador: str, coluna: str) -> float:
            return float(kpis.loc[kpis["Indicador"] == indicador, coluna].iloc[0])

        self.titulo("DRE 2019 - RESUMO EXECUTIVO SENIOR")

        receita_real = valor("Receita Bruta", "Realizado")
        receita_orc = valor("Receita Bruta", "Orçado")
        lucro_real = valor("Lucro Líquido", "Realizado")
        lucro_orc = valor("Lucro Líquido", "Orçado")
        lucro_desvio = valor("Lucro Líquido", "Desvio")
        lucro_desvio_perc = valor("Lucro Líquido", "Desvio_%")
        margem_real = valor("Margem Líquida", "Realizado") * 100
        margem_orc = valor("Margem Líquida", "Orçado") * 100
        margem_desvio = margem_real - margem_orc

        print(f"Receita Bruta Realizada:  {self.moeda(receita_real, compacto=True)}")
        print(f"Receita Bruta Orçada:     {self.moeda(receita_orc, compacto=True)}")
        print(f"Desvio Receita:           {self.moeda(receita_real - receita_orc, compacto=True)} ({self.percentual((receita_real - receita_orc) / receita_orc * 100)})")
        print()
        print(f"Lucro Líquido Realizado:  {self.moeda(lucro_real, compacto=True)}")
        print(f"Lucro Líquido Orçado:     {self.moeda(lucro_orc, compacto=True)}")
        print(f"Gap de Resultado:         {self.moeda(lucro_desvio, compacto=True)} ({self.percentual(lucro_desvio_perc)})")
        print()
        print(f"Margem Líquida Realizada: {self.percentual(margem_real)}")
        print(f"Margem Líquida Orçada:    {self.percentual(margem_orc)}")
        print(f"Gap Margem Líquida:       {self.percentual(margem_desvio, pp=True)}")
        print()
        print(f"Receita acima da meta:    {int(receita['Meta_Batida'].sum())}/{len(receita)} meses")
        print(f"Lucro acima da meta:      {int(lucro_mensal['Meta_Batida'].sum())}/{len(lucro_mensal)} meses")

        print()
        print("Maiores drivers do gap de resultado:")
        for indice, linha in enumerate(top_desvios.itertuples(index=False), start=1):
            nivel_2 = getattr(linha, self.config.coluna_nivel_2.replace(" ", "_"), linha[1])
            print(f"{indice}. {linha.Status} | {linha.Favorabilidade} | {nivel_2}: {self.moeda(linha.Impacto_Resultado, compacto=True)}")

        print()
        if lucro_desvio < 0:
            print("Conclusão: resultado anual abaixo do orçamento; a prioridade é atacar os drivers desfavoráveis de receita e custo.")
        else:
            print("Conclusão: resultado anual acima do orçamento; a prioridade é proteger os drivers favoráveis e investigar recorrência.")

    def mostrar_kpis_dre(self) -> None:
        self.titulo("KPIs DA DRE - REALIZADO VS ORÇADO")

        kpis = self.calcular_kpis()
        linhas_formatadas = []
        for _, row in kpis.iterrows():
            eh_percentual = row["Indicador"] in self.indicadores_percentuais
            linhas_formatadas.append(
                {
                    "Indicador": row["Indicador"],
                    "Realizado": self.percentual(row["Realizado"] * 100) if eh_percentual else self.moeda(row["Realizado"]),
                    "Orçado": self.percentual(row["Orçado"] * 100) if eh_percentual else self.moeda(row["Orçado"]),
                    "Desvio": self.percentual(row["Desvio"] * 100, pp=True) if eh_percentual else self.moeda(row["Desvio"]),
                    "Desvio_%": "-" if eh_percentual else self.percentual(row["Desvio_%"]),
                    "Favorabilidade": row["Favorabilidade"],
                    "Status": row["Status"],
                }
            )

        print(pd.DataFrame(linhas_formatadas).to_string(index=False))

    def mostrar_top_desvios(self) -> None:
        self.titulo("TOP 10 DRIVERS DE DESVIO - NÍVEL 2")

        tabela = self.resumo_nivel_2().head(10)
        tabela = tabela[
            [
                self.config.coluna_status,
                self.config.coluna_favorabilidade,
                self.config.coluna_nivel_1,
                self.config.coluna_nivel_2,
                "Realizado",
                "Orçado",
                "Desvio",
                "Impacto_Resultado",
                "Materialidade_%_Lucro_Orçado",
            ]
        ]
        tabela = self.formatar_tabela_moeda(tabela, ["Realizado", "Orçado", "Desvio", "Impacto_Resultado"], ["Materialidade_%_Lucro_Orçado"])
        print(tabela.to_string(index=False))

    def mostrar_evolucao_mensal(self) -> None:
        self.titulo("EVOLUÇÃO MENSAL - RECEITA, LUCRO E GAP ACUMULADO")

        receita = self.resumo_receita_mensal()
        lucro = self.resumo_lucro_liquido_mensal()
        tabela = receita[["Mês", "Receita_Realizada", "Receita_Orçada", "Desvio", "Desvio_%", "Meta_Batida"]].merge(
            lucro[["Mês", "Lucro_Realizado", "Lucro_Orçado", "Desvio", "Desvio_Acumulado", "Meta_Batida"]],
            on="Mês",
            suffixes=("_Receita", "_Lucro"),
        )

        tabela["Meta_Batida_Receita"] = tabela["Meta_Batida_Receita"].map({True: "Sim", False: "Não"})
        tabela["Meta_Batida_Lucro"] = tabela["Meta_Batida_Lucro"].map({True: "Sim", False: "Não"})

        tabela = self.formatar_tabela_moeda(
            tabela,
            ["Receita_Realizada", "Receita_Orçada", "Desvio_Receita", "Lucro_Realizado", "Lucro_Orçado", "Desvio_Lucro", "Desvio_Acumulado"],
            ["Desvio_%"],
        )
        print(tabela.to_string(index=False))

    def mostrar_detalhamento_nivel_1(self) -> None:
        self.titulo("DETALHAMENTO POR NÍVEL 1")

        tabela = self.resumo_nivel_1()
        tabela = tabela[
            [
                self.config.coluna_status,
                self.config.coluna_favorabilidade,
                self.config.coluna_nivel_1,
                "Realizado",
                "Orçado",
                "Desvio",
                "Impacto_Resultado",
                "Materialidade_%_Lucro_Orçado",
                "Contas",
            ]
        ]
        tabela = self.formatar_tabela_moeda(tabela, ["Realizado", "Orçado", "Desvio", "Impacto_Resultado"], ["Materialidade_%_Lucro_Orçado"])
        print(tabela.to_string(index=False))

    def mostrar_detalhamento_nivel_2(self) -> None:
        self.titulo("DETALHAMENTO POR NÍVEL 2")

        tabela = self.resumo_nivel_2()
        tabela = tabela[
            [
                self.config.coluna_status,
                self.config.coluna_favorabilidade,
                self.config.coluna_nivel_1,
                self.config.coluna_nivel_2,
                "Realizado",
                "Orçado",
                "Desvio",
                "Impacto_Resultado",
                "Materialidade_%_Lucro_Orçado",
                "Contas",
            ]
        ]
        tabela = self.formatar_tabela_moeda(tabela, ["Realizado", "Orçado", "Desvio", "Impacto_Resultado"], ["Materialidade_%_Lucro_Orçado"])
        print(tabela.to_string(index=False))

    def mostrar_auditoria(self) -> None:
        self.titulo("AUDITORIA DOS DADOS")
        print(self.auditoria_dados().to_string(index=False))

    def aplicar_estilo_planilha(self, writer: pd.ExcelWriter) -> None:
        workbook = writer.book
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        ok_fill = PatternFill("solid", fgColor="E2F0D9")
        alerta_fill = PatternFill("solid", fgColor="FFF2CC")
        critico_fill = PatternFill("solid", fgColor="F4CCCC")
        favoravel_fill = PatternFill("solid", fgColor="D9EAD3")
        desfavoravel_fill = PatternFill("solid", fgColor="FCE4D6")
        border = Border(bottom=Side(style="thin", color="D9E2F3"))

        for ws in workbook.worksheets:
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                cell.border = border

            headers = {cell.value: cell.column for cell in ws[1]}
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.border = border
                    cell.alignment = Alignment(vertical="center")

            for nome_coluna, col_idx in headers.items():
                letra = get_column_letter(col_idx)
                valores = [ws.cell(row=linha, column=col_idx).value for linha in range(1, ws.max_row + 1)]
                largura = min(max(len(str(valor)) if valor is not None else 0 for valor in valores) + 2, 38)
                ws.column_dimensions[letra].width = max(largura, 12)

                if nome_coluna and ("Valor" in str(nome_coluna) or "Realizado" in str(nome_coluna) or "Orçado" in str(nome_coluna) or "Desvio" in str(nome_coluna) or "Impacto" in str(nome_coluna)):
                    for cell in ws[letra][1:]:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = 'R$ #,##0;[Red]-R$ #,##0'
                if nome_coluna and ("%_" in str(nome_coluna) or str(nome_coluna).endswith("_%") or str(nome_coluna) == "Pareto_%"):
                    for cell in ws[letra][1:]:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '0.00"%"'
                if nome_coluna == "Mês/Ano":
                    for cell in ws[letra][1:]:
                        cell.number_format = "mmm/yy"

            if "Status" in headers:
                letra = get_column_letter(headers["Status"])
                intervalo = f"{letra}2:{letra}{ws.max_row}"
                ws.conditional_formatting.add(intervalo, CellIsRule(operator="equal", formula=['"OK"'], fill=ok_fill))
                ws.conditional_formatting.add(intervalo, CellIsRule(operator="equal", formula=['"ATENÇÃO"'], fill=alerta_fill))
                ws.conditional_formatting.add(intervalo, CellIsRule(operator="equal", formula=['"CRÍTICO"'], fill=critico_fill))

            if "Favorabilidade" in headers:
                letra = get_column_letter(headers["Favorabilidade"])
                intervalo = f"{letra}2:{letra}{ws.max_row}"
                ws.conditional_formatting.add(intervalo, CellIsRule(operator="equal", formula=['"Favorável"'], fill=favoravel_fill))
                ws.conditional_formatting.add(intervalo, CellIsRule(operator="equal", formula=['"Desfavorável"'], fill=desfavoravel_fill))

            if ws.title == "KPIs" and "Indicador" in headers:
                col_indicador = headers["Indicador"]
                for linha in range(2, ws.max_row + 1):
                    indicador = ws.cell(row=linha, column=col_indicador).value
                    if indicador in self.indicadores_percentuais:
                        for nome_coluna in ["Realizado", "Orçado"]:
                            if nome_coluna in headers:
                                ws.cell(row=linha, column=headers[nome_coluna]).number_format = "0.00%"
                        if "Desvio" in headers:
                            ws.cell(row=linha, column=headers["Desvio"]).number_format = '0.00" p.p."'

            if ws.title == "Resumo Executivo" and {"Indicador", "Valor"}.issubset(headers):
                col_indicador = headers["Indicador"]
                col_valor = headers["Valor"]
                for linha in range(2, ws.max_row + 1):
                    indicador = str(ws.cell(row=linha, column=col_indicador).value)
                    celula_valor = ws.cell(row=linha, column=col_valor)
                    if "Margem" in indicador:
                        celula_valor.number_format = "0.00%"
                    elif indicador.startswith("Meses"):
                        celula_valor.number_format = "0"

        if "Resumo Executivo" in workbook.sheetnames:
            workbook["Resumo Executivo"].freeze_panes = "A2"

    def exportar_tabelas(self) -> None:
        self.pasta_saidas.mkdir(exist_ok=True)
        caminho_saida = self.pasta_saidas / ARQUIVO_SAIDA

        with pd.ExcelWriter(caminho_saida, engine="openpyxl") as writer:
            self.resumo_executivo_df().to_excel(writer, sheet_name="Resumo Executivo", index=False)
            self.calcular_dre_estruturada().to_excel(writer, sheet_name="DRE Estruturada", index=False)
            self.calcular_kpis().to_excel(writer, sheet_name="KPIs", index=False)
            self.resumo_nivel_1().to_excel(writer, sheet_name="Nivel 1", index=False)
            self.resumo_nivel_2().to_excel(writer, sheet_name="Nivel 2", index=False)
            self.resumo_receita_mensal().to_excel(writer, sheet_name="Receita Mensal", index=False)
            self.resumo_lucro_liquido_mensal().to_excel(writer, sheet_name="Lucro Mensal", index=False)
            self.contribuicao_gap_lucro().to_excel(writer, sheet_name="Gap Lucro", index=False)
            self.pareto_desvios().to_excel(writer, sheet_name="Pareto Desvios", index=False)
            self.melhores_piores_meses().to_excel(writer, sheet_name="Melhores Piores Meses", index=False)
            self.auditoria_dados().to_excel(writer, sheet_name="Auditoria", index=False)
            self.df_final.to_excel(writer, sheet_name="Base Consolidada", index=False)
            self.aplicar_estilo_planilha(writer)

    def configurar_graficos(self) -> None:
        self.pasta_graficos.mkdir(exist_ok=True)
        sns.set_theme(style="whitegrid", palette="deep")
        plt.rcParams["figure.dpi"] = 120
        plt.rcParams["savefig.dpi"] = 300
        plt.rcParams["axes.titlesize"] = 14
        plt.rcParams["axes.labelsize"] = 10
        plt.rcParams["axes.formatter.useoffset"] = False

    def salvar_grafico(self, nome: str) -> None:
        plt.tight_layout()
        plt.savefig(self.pasta_graficos / nome, bbox_inches="tight")
        plt.close()

    def grafico_kpis(self) -> None:
        kpis = self.calcular_kpis()
        base = kpis[kpis["Indicador"].isin(["Receita Bruta", "Receita Líquida", "Lucro Bruto", "Lucro Líquido"])].copy()
        melt = base.melt(id_vars="Indicador", value_vars=["Orçado", "Realizado"], var_name="Cenário", value_name="Valor")

        plt.figure(figsize=(12, 6))
        ax = sns.barplot(data=melt, x="Indicador", y="Valor", hue="Cenário", palette=[self.cor_orcado, self.cor_realizado])
        ax.set_title("KPIs Principais da DRE - Orçado vs Realizado")
        ax.set_xlabel("")
        ax.set_ylabel("Valor anual")
        ax.yaxis.set_major_formatter(FuncFormatter(self.formatar_milhoes_eixo))
        ax.tick_params(axis="x", rotation=10)

        for container in ax.containers:
            ax.bar_label(container, labels=[self.moeda(v.get_height(), compacto=True) for v in container], fontsize=8, padding=3)

        self.salvar_grafico("01_kpis_principais_dre.png")

    def grafico_top_desvios(self) -> None:
        top = self.resumo_nivel_2().head(10).copy().sort_values("Impacto_Resultado")
        top["Categoria"] = top[self.config.coluna_nivel_2].astype(str)
        cores = [self.cor_favoravel if valor > 0 else self.cor_desfavoravel for valor in top["Impacto_Resultado"]]

        fig, ax = plt.subplots(figsize=(13, 7))
        ax.barh(top["Categoria"], top["Impacto_Resultado"], color=cores)
        ax.set_title("Top 10 Desvios por Impacto no Resultado - Nível 2")
        ax.set_xlabel("Impacto Realizado - Orçado")
        ax.set_ylabel("")
        ax.xaxis.set_major_formatter(FuncFormatter(self.formatar_milhoes_eixo))
        ax.axvline(0, color="#263238", linewidth=1)

        limite = max(abs(top["Impacto_Resultado"].min()), abs(top["Impacto_Resultado"].max()))
        ax.set_xlim(-limite * 1.28, limite * 1.28)
        deslocamento = limite * 0.035
        for y, valor in enumerate(top["Impacto_Resultado"]):
            x = valor + deslocamento if valor >= 0 else valor - deslocamento
            ha = "left" if valor >= 0 else "right"
            ax.text(x, y, self.moeda(valor, compacto=True), va="center", ha=ha, fontsize=9)

        self.salvar_grafico("02_top_10_desvios_nivel_2.png")

    def grafico_receita_mensal(self) -> None:
        receita = self.resumo_receita_mensal()
        melt = receita.melt(
            id_vars=[self.config.coluna_mes_ano, "Mês"],
            value_vars=["Receita_Orçada", "Receita_Realizada"],
            var_name="Cenário",
            value_name="Valor",
        )

        plt.figure(figsize=(13, 6))
        ax = sns.lineplot(data=melt, x=self.config.coluna_mes_ano, y="Valor", hue="Cenário", marker="o", linewidth=2.5, palette=[self.cor_orcado, self.cor_realizado])
        ax.set_title("Receita Bruta Mensal - Meta vs Execução")
        ax.set_xlabel("Mês")
        ax.set_ylabel("Receita Bruta")
        ax.yaxis.set_major_formatter(FuncFormatter(self.formatar_milhoes_eixo))
        ax.tick_params(axis="x", rotation=45)

        for _, row in receita.iterrows():
            cor = self.cor_favoravel if row["Meta_Batida"] else self.cor_desfavoravel
            ax.scatter(row[self.config.coluna_mes_ano], row["Receita_Realizada"], color=cor, s=55, zorder=5)

        self.salvar_grafico("03_receita_bruta_mensal.png")

    def grafico_lucro_mensal(self) -> None:
        lucro = self.resumo_lucro_liquido_mensal()
        melt = lucro.melt(
            id_vars=[self.config.coluna_mes_ano, "Mês"],
            value_vars=["Lucro_Orçado", "Lucro_Realizado"],
            var_name="Cenário",
            value_name="Valor",
        )

        fig, ax = plt.subplots(figsize=(13, 6))
        sns.lineplot(data=melt, x=self.config.coluna_mes_ano, y="Valor", hue="Cenário", marker="o", linewidth=2.5, palette=[self.cor_orcado, self.cor_realizado], ax=ax)
        ax.set_title("Lucro Líquido Mensal - Orçado vs Realizado")
        ax.set_xlabel("Mês")
        ax.set_ylabel("Lucro Líquido")
        ax.yaxis.set_major_formatter(FuncFormatter(self.formatar_milhoes_eixo))
        ax.tick_params(axis="x", rotation=45)

        ax2 = ax.twinx()
        ax2.plot(lucro[self.config.coluna_mes_ano], lucro["Desvio_Acumulado"], color="#6A1B9A", linestyle="--", linewidth=2, label="Gap acumulado")
        ax2.yaxis.set_major_formatter(FuncFormatter(self.formatar_milhoes_eixo))
        ax2.set_ylabel("Gap acumulado")
        ax2.grid(False)
        linhas, labels = ax.get_legend_handles_labels()
        linhas2, labels2 = ax2.get_legend_handles_labels()
        ax.legend(linhas + linhas2, labels + labels2, loc="best")

        self.salvar_grafico("04_lucro_liquido_mensal.png")

    def grafico_waterfall_resultado(self) -> None:
        contribuicao = self.contribuicao_gap_lucro()
        dre = self.calcular_dre_estruturada()
        lucro_orcado = self.valor_indicador(dre, "Lucro Líquido", "Orçado")
        lucro_realizado = self.valor_indicador(dre, "Lucro Líquido", "Realizado")

        labels = ["Lucro Orçado"] + contribuicao["Categoria"].tolist() + ["Lucro Realizado"]
        valores = [lucro_orcado] + contribuicao["Impacto_Resultado"].tolist() + [lucro_realizado]
        cores = [self.cor_neutra] + [self.cor_favoravel if v >= 0 else self.cor_desfavoravel for v in contribuicao["Impacto_Resultado"]] + [self.cor_realizado]

        bottoms = [0.0]
        heights = [lucro_orcado]
        acumulado = lucro_orcado
        pontos = [lucro_orcado]
        for impacto in contribuicao["Impacto_Resultado"]:
            novo = acumulado + impacto
            bottoms.append(min(acumulado, novo))
            heights.append(abs(impacto))
            acumulado = novo
            pontos.append(acumulado)
        bottoms.append(0.0)
        heights.append(lucro_realizado)
        pontos.append(lucro_realizado)

        fig, ax = plt.subplots(figsize=(13, 6))
        x = range(len(labels))
        ax.bar(x, heights, bottom=bottoms, color=cores, width=0.62)
        ax.set_title("Waterfall do Resultado - Lucro Orçado até Lucro Realizado")
        ax.set_ylabel("Lucro Líquido")
        ax.yaxis.set_major_formatter(FuncFormatter(self.formatar_milhoes_eixo))
        ax.set_xticks(list(x))
        ax.set_xticklabels(labels, rotation=20, ha="right")

        margem = max(abs(max(pontos) - min(pontos)) * 0.30, 2_000_000)
        ax.set_ylim(min(pontos) - margem, max(pontos) + margem)
        for i, (bottom, height, valor) in enumerate(zip(bottoms, heights, valores)):
            y = bottom + height
            texto = self.moeda(valor, compacto=True) if i in (0, len(labels) - 1) else self.moeda(valor, compacto=True)
            ax.text(i, y, texto, ha="center", va="bottom", fontsize=8)

        for i in range(1, len(labels) - 1):
            ax.plot([i - 0.31, i + 0.31], [pontos[i], pontos[i]], color="#455A64", linewidth=0.8)

        self.salvar_grafico("05_ponte_impacto_resultado.png")

    def grafico_pareto_desvios(self) -> None:
        pareto = self.pareto_desvios().head(12).copy()
        pareto["Categoria"] = pareto[self.config.coluna_nivel_2].astype(str).str.replace(r"^\d+\s-\s", "", regex=True)

        fig, ax = plt.subplots(figsize=(13, 6))
        ax.bar(pareto["Categoria"], pareto["Impacto_Absoluto"], color="#455A64")
        ax.set_title("Pareto dos Desvios - Materialidade por Nível 2")
        ax.set_ylabel("Impacto absoluto")
        ax.yaxis.set_major_formatter(FuncFormatter(self.formatar_milhoes_eixo))
        ax.tick_params(axis="x", rotation=35)

        ax2 = ax.twinx()
        ax2.plot(pareto["Categoria"], pareto["Pareto_%"], color="#C62828", marker="o", linewidth=2)
        ax2.axhline(80, color="#C62828", linestyle="--", linewidth=1)
        ax2.set_ylim(0, 105)
        ax2.set_ylabel("Acumulado")
        ax2.yaxis.set_major_formatter(FuncFormatter(lambda valor, _pos: f"{valor:.0f}%"))
        ax2.grid(False)

        self.salvar_grafico("06_pareto_desvios.png")

    def gerar_graficos(self) -> None:
        self.titulo("GRÁFICOS GERADOS")
        self.configurar_graficos()
        self.grafico_kpis()
        self.grafico_top_desvios()
        self.grafico_receita_mensal()
        self.grafico_lucro_mensal()
        self.grafico_waterfall_resultado()
        self.grafico_pareto_desvios()

        for nome in [
            "01_kpis_principais_dre.png",
            "02_top_10_desvios_nivel_2.png",
            "03_receita_bruta_mensal.png",
            "04_lucro_liquido_mensal.png",
            "05_ponte_impacto_resultado.png",
            "06_pareto_desvios.png",
        ]:
            print(nome)
        print(f"Pasta: {self.pasta_graficos.resolve()}")

    def mostrar_reflexao_exercicio(self) -> None:
        self.titulo("AÇÃO E REFLEXÃO - RESPOSTA AO ENUNCIADO")

        print("Problemas de negócio e desafios técnicos identificados:")
        print("1. A DRE precisa ser governada pela estrutura oficial da aba Formato DRE, não por regras soltas no código.")
        print("2. Realizado vs Orçado só é confiável depois de validar chaves, contas, datas, orçamento zero e mapeamento contábil.")
        print("3. Desvio contábil, impacto no resultado e favorabilidade são conceitos diferentes e agora aparecem separados.")
        print("4. A priorização executiva deve combinar valor absoluto, materialidade sobre receita/lucro e recorrência mensal.")
        print()
        print("Insight estratégico:")
        print("O lucro ficou abaixo do orçamento principalmente por pressão de receita e custo; o Pareto mostra onde agir primeiro.")

    def executar(self) -> None:
        self.carregar_abas()
        self.validar_colunas_obrigatorias()
        self.limpar_dados()
        self.consolidar_bases()

        self.mostrar_resumo_executivo()
        self.mostrar_kpis_dre()
        self.mostrar_top_desvios()
        self.mostrar_evolucao_mensal()
        self.mostrar_detalhamento_nivel_1()
        self.mostrar_detalhamento_nivel_2()
        self.mostrar_reflexao_exercicio()
        self.mostrar_auditoria()
        self.exportar_tabelas()
        self.gerar_graficos()

        self.titulo("PROCESSO FINALIZADO")
        print(f"Análise concluída em {time.perf_counter() - self.inicio:.2f} segundos.")
        print(f"Arquivo Excel de saída: {(self.pasta_saidas / ARQUIVO_SAIDA).resolve()}")


def verificar_dependencias() -> None:
    pacotes = {
        "pandas": "pandas",
        "openpyxl": "openpyxl",
        "matplotlib": "matplotlib",
        "seaborn": "seaborn",
    }

    faltando = []
    for modulo, pacote in pacotes.items():
        try:
            __import__(modulo)
        except ModuleNotFoundError:
            faltando.append(pacote)

    if faltando:
        print("Bibliotecas faltando:")
        for pacote in faltando:
            print(f"- {pacote}")
        print()
        print("Instale com:")
        print(f"pip install {' '.join(faltando)}")
        sys.exit(1)


def criar_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Análise senior de DRE 2019: Realizado vs Orçado.")
    parser.add_argument("--arquivo", default=ARQUIVO_EXCEL, help="Caminho da planilha de entrada.")
    parser.add_argument("--pasta-saidas", default=PASTA_SAIDAS, help="Pasta para o Excel final.")
    parser.add_argument("--pasta-graficos", default=PASTA_GRAFICOS, help="Pasta para os gráficos PNG.")
    return parser


def main(argv: Sequence[str] | None = None) -> None:
    verificar_dependencias()
    args = criar_parser().parse_args(argv)
    analise = AnaliseDRESenior(
        arquivo_excel=args.arquivo,
        pasta_saidas=args.pasta_saidas,
        pasta_graficos=args.pasta_graficos,
    )
    analise.executar()


if __name__ == "__main__":
    main()
